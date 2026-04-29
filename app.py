import sys
import os
import pdfplumber
import pandas as pd
from version import APP_NAME
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QTextEdit, QFileDialog, QMessageBox,
    QInputDialog, QLineEdit
)
from PySide6.QtCore import Qt

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.resize(700, 500)

        # Central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        # Sidebar setup
        sidebar = QWidget()
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar.setFixedWidth(200)

        self.logo_label = QLabel(APP_NAME)
        self.logo_label.setAlignment(Qt.AlignCenter)
        self.logo_label.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 20px;")
        sidebar_layout.addWidget(self.logo_label)

        self.select_file_btn = QPushButton("Select File")
        self.select_file_btn.clicked.connect(self.select_file)
        sidebar_layout.addWidget(self.select_file_btn)

        self.extract_btn = QPushButton("Extract Data")
        self.extract_btn.clicked.connect(self.extract_data)
        self.extract_btn.setEnabled(False)
        sidebar_layout.addWidget(self.extract_btn)

        sidebar_layout.addStretch()
        main_layout.addWidget(sidebar)

        # Main content setup
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)

        self.file_label = QLabel("Selected: No file selected")
        self.file_label.setStyleSheet("font-size: 14px; margin-bottom: 10px;")
        content_layout.addWidget(self.file_label)

        self.log_textbox = QTextEdit()
        self.log_textbox.setReadOnly(True)
        content_layout.addWidget(self.log_textbox)

        main_layout.addWidget(content_widget)

        self.pdf_file = ""

    def log(self, message):
        self.log_textbox.append(message)
        # Force UI update
        QApplication.processEvents()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select PDF", "", "PDF Files (*.pdf)")
        if file_path:
            self.set_file(file_path)

    def set_file(self, file_path):
        self.pdf_file = file_path
        self.file_label.setText(f"Selected: {os.path.basename(self.pdf_file)}")
        self.log(f"Selected file: {self.pdf_file}")
        
        if self.pdf_file:
            self.extract_btn.setEnabled(True)
        else:
            self.extract_btn.setEnabled(False)

    def extract_data(self):
        if not self.pdf_file:
            return
            
        self.log("\nStarting extraction...")
        self.extract_btn.setEnabled(False)
        self.select_file_btn.setEnabled(False)
        
        output_folder = os.path.join(os.path.dirname(self.pdf_file), "extracted_output")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            
        pdf_file = os.path.basename(self.pdf_file)
        pdf_path = self.pdf_file
        self.log(f"Processing {pdf_file}...")
        
        password = None
        # Check if PDF is encrypted and ask for password if needed
        try:
            with pdfplumber.open(pdf_path) as test_pdf:
                pass
        except Exception as e:
            error_msg = str(e).lower()
            error_type = type(e).__name__.lower()
            # If the error message or error type contains any hint of encryption or pdfminer issues
            if any(kw in error_msg for kw in ["password", "authenticate", "pdfminer", "encrypted"]) or \
               any(kw in error_type for kw in ["password", "pdfminer", "crypt"]):
                while True:
                    text, ok = QInputDialog.getText(
                        self, "Password Required", 
                        f"The file '{pdf_file}' is password protected.\nPlease enter the password:",
                        QLineEdit.Password
                    )
                    if ok:
                        try:
                            with pdfplumber.open(pdf_path, password=text) as test_pdf:
                                password = text
                                break
                        except:
                            QMessageBox.warning(self, "Error", "Incorrect password. Please try again.")
                    else:
                        self.log("Extraction cancelled (password not provided).")
                        self.extract_btn.setEnabled(True)
                        self.select_file_btn.setEnabled(True)
                        return
            else:
                self.log(f"  -> Info: PDF check failed with {type(e).__name__}: {str(e)}")
                # Continue anyway, it might be a different issue caught later
        
        def save_styled_excel(dataframe, path):
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                dataframe.to_excel(writer, index=False, header=False, sheet_name='Transactions')
                worksheet = writer.sheets['Transactions']
                
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        try:
            import re
            
            # Support multiple date formats:
            # 1. 01-01-2025 or 01/01/25
            # 2. 04JUN25 (Emirates NBD)
            # 3. 30 Apr, 2025 (Wio)
            date_pattern = re.compile(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{1,2}[A-Z]{3}\d{2}|\d{1,2}\s+[A-Za-z]{3},\s+\d{4}')
            
            skip_keywords = [
                'your bank statement', 'account summary', 'opening balance',
                'closing balance', 'total deposits', 'total withdrawals',
                'interest rate', 'overdue charges', 'interest type',
                'please see', 'important information', 'the national bank',
                '"rakbank"', 'central bank of', 'page [', 'page ', 'disclaimer',
                'end of day', 'summary ', 'payments in', 'payments out',
                'get in touch', 'for queries', 'for complaints', 'call 04',
                'we aim', 'this is a digitally', 'accrued interest',
                'account number:', 'branch:', 'currency:', 'iban:',
                'account type:', 'statement period:', 'date issued:',
                '24hr customer service', 'www.starling', 'sort code:',
                'account name:', 'financial services', 'your deposit',
                '© 2025', 'wio bank', 'this bank is regulated',
                'please review', 'within 30 days', 'correct (subject',
                'to raise a complaint', 'summary of accounts',
                'account statement', 'account holder',
                'the items and balance', 'verified. report any',
                'of the statement date', 'accurate.', 'all charges, terms',
                'please note that for foreign', 'indicative only',
                'mashreqbank psc', 'it is the policy', 'full compliance',
                'with iran, syria', 'dear customer', 'statement for period',
                'customer number', 'account currency', 'account number ',
                'يكنبلا', 'يراجلا', 'خيراتلا', 'فصولا', 'كيشلا',
                'بحسلا', 'ةعيدولا', 'ديصرلا', 'باسحلا', 'ةدئافلا',
                'ةرخأتملا', 'ةفاضملا', '.ةمهم', 'ينطولا', 'يزكرملا',
                'تامولعم', 'فشكلا', 'رادصإلا',
                'end of day', 'account balance', 'carried forward',
                'is registered in england', 'prudential regulation authority',
                'financial conduct authority', 'london fruit and wool exchange',
                'fscs website', 'information sheet and exclusions',
                'authorised by the', 'regulated by the', 'starling bank limited',
                'duval square', 'e1 6pw', 'number 730166',
                'list which are', 'available in the app and', 'on our website.',
                'further details can also be found', 'refer to the fscs website',
                'we charge', 'interest each day you', 'date range applicable',
                'unlimited 0.00%', 'less than', 'interest rates', 'overdraft',
                'united arab', 'emirates p.o box', 'digital stamp',
                'does not require signature', 'description (incl. vat)',
                'amount balance', 'ref. number', 'tax registration number',
                'brought forward', 'carried forward', 'need help?',
                'following channels', 'personal banking', 'business banking',
                'private banking', 'dedicated relationship', 'online banking',
                'future reference', 'dispute resolution', 'sanadak.ae',
                'centralbank.ae', 'complaint with the bank', 'contact us',
                'customer service', 'registered details:', 'paid up capital',
                'commercial registration', 'head office:', 'p.j.s.c',
                'licensed by the central bank', 'emirates nbd bank',
                'phone banking:', 'email at', 'complaint-process',
                'customersupport@', 'www.emiratesnbd.com', 'visiting any',
                'nbd branch', 'emirates nbd branch', '800 54', '800456',
                'p.o. box', 'uae', 'date description', 'debits credits',
                'credits balance', 'description debits', 'end of statement',
                'total debits', 'total credits', 'total amount', 'total',
            ]
            
            def is_skip_line(line):
                lower = line.lower().strip()
                if not lower:
                    return True
                
                # 1. Skip if any keyword matches
                for kw in skip_keywords:
                    if kw in lower:
                        return True
                
                # 2. Skip ANY line containing Arabic characters (not used for data)
                if any('\u0600' <= c <= '\u06FF' for c in line):
                    return True
                
                # 3. Skip page numbers (e.g., "page 1 of 4", "page1")
                if re.search(r'page\s*\d+', lower):
                    return True
                
                # 4. Skip lines that have no English letters and no digits (junk punctuation)
                letters = sum(1 for c in lower if 'a' <= c <= 'z')
                digits = sum(1 for c in lower if '0' <= c <= '9')
                if letters == 0 and digits == 0 and len(lower) > 2:
                    return True
                    
                # 5. Skip if it looks like a phone number or URL
                if re.search(r'\(\+\d+\)|\d{3}-\d{4}|www\.|http', lower):
                    return True
                
                # 6. Skip if the line is JUST a header word or common footer word
                if lower in ['date', 'description', 'debits', 'credits', 'balance', 'total']:
                    return True
                    
                return False
            
            # Step 1: Extract text and find header line
            all_lines = []
            with pdfplumber.open(pdf_path, password=password) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        for line in text.split('\n'):
                            all_lines.append(line.strip())
            
            if not all_lines:
                self.log(f"  -> No text found in {pdf_file}.")
                self.extract_btn.setEnabled(True)
                self.select_file_btn.setEnabled(True)
                return
            
            header_kws_detect = ['date', 'description', 'balance', 'amount', 'withdrawal', 
                                 'deposit', 'credit', 'debit', 'transaction', 'type', 'ref.', 'reference', 
                                 'debits', 'credits', 'value date', 'chq/ref', 'currency']
            header_line = None
            for i, line in enumerate(all_lines):
                lower = line.lower().strip()
                if 'date' not in lower:
                    continue
                matches = sum(1 for kw in header_kws_detect if kw in lower)
                if matches >= 2:
                    header_line = line.strip()
                    break
            
            if not header_line:
                self.log(f"  -> Could not detect table header in {pdf_file}.")
                self.extract_btn.setEnabled(True)
                self.select_file_btn.setEnabled(True)
                return
            
            # Step 2: Use word-level extraction to find header column positions
            known_multi_word = {
                'reference number': 'Reference Number',
                'account balance': 'Account Balance',
                'ref. number': 'Ref. Number',
                'payments in': 'Payments In',
                'payments out': 'Payments Out',
                'value date': 'Value Date',
                'chq/ref no.': 'Chq/Ref No.',
                'chq/ref': 'Chq/Ref No.',
                'running balance': 'Running Balance',
            }
            
            with pdfplumber.open(pdf_path, password=password) as pdf:
                # Find header words on the page
                header_words_found = []
                header_top = None
                header_page_idx = None
                
                for page_idx, page in enumerate(pdf.pages):
                    words = page.extract_words()
                    # Look for a line with at least 3 header keywords
                    # Group words by y position
                    y_groups = {}
                    for w in words:
                        y_key = round(w['top'] / 4) * 4
                        if y_key not in y_groups:
                            y_groups[y_key] = []
                        y_groups[y_key].append(w)
                    
                    sorted_y_keys = sorted(y_groups.keys())
                    for y_idx, y_key in enumerate(sorted_y_keys):
                        line_words = y_groups[y_key]
                        line_text = ' '.join([w['text'].lower() for w in line_words])
                        kw_matches = sum(1 for kw in header_kws_detect if kw in line_text)
                        if kw_matches >= 3 and 'date' in line_text:
                            header_top = line_words[0]['top']
                            header_page_idx = page_idx
                            
                            # Collect header words from this line AND adjacent lines (±10 y)
                            for adj_y in sorted_y_keys:
                                if abs(adj_y - y_key) <= 12:  # Within ~12 px
                                    for w in y_groups[adj_y]:
                                        wt = w['text'].lower()
                                        if wt in header_kws_detect or wt in ('in', 'out', 'number', 'cheque', 'ref.', 'reference', 'account', 'no.', 'value', 'running'):
                                            header_words_found.append(w)
                                    # Update header_top to the latest line if it's after
                                    if adj_y > y_key:
                                        header_top = max(header_top, max(w['top'] for w in y_groups[adj_y]))
                            break
                    if header_words_found:
                        break
                
                if not header_words_found:
                    self.log(f"  -> Could not locate header positions in {pdf_file}.")
                    self.extract_btn.setEnabled(True)
                    self.select_file_btn.setEnabled(True)
                    return
                
                # Sort by x position and merge multi-word columns
                header_words_found.sort(key=lambda w: w['x0'])
                
                merged_cols = []
                i = 0
                while i < len(header_words_found):
                    w = header_words_found[i]
                    # Skip Arabic words
                    if any(ord(c) > 127 for c in w['text']):
                        i += 1
                        continue
                        
                    # Try 3-word match
                    if i + 2 < len(header_words_found):
                        nw1 = header_words_found[i + 1]
                        nw2 = header_words_found[i + 2]
                        three = (w['text'] + ' ' + nw1['text'] + ' ' + nw2['text']).lower()
                        if three in known_multi_word:
                            merged_cols.append({'text': known_multi_word[three], 'x0': w['x0'], 'x1': nw2['x1']})
                            i += 3
                            continue
                            
                    # Try 2-word match
                    if i + 1 < len(header_words_found):
                        nw = header_words_found[i + 1]
                        two = (w['text'] + ' ' + nw['text']).lower()
                        if two in known_multi_word:
                            merged_cols.append({
                                'text': known_multi_word[two],
                                'x0': w['x0'], 'x1': nw['x1']
                            })
                            i += 2
                            continue
                    merged_cols.append({'text': w['text'].capitalize(), 'x0': w['x0'], 'x1': w['x1']})
                    i += 1
                
                col_names = [c['text'] for c in merged_cols]
                
                # Build column boundaries (bias towards the end of the current column to avoid wide gaps absorbing next column data)
                col_ranges = []
                for ci, col in enumerate(merged_cols):
                    if ci == 0:
                        x_start = 0
                    else:
                        # Start of this column is influenced by the previous column's end
                        prev_col = merged_cols[ci - 1]
                        x_start = prev_col['x1'] + (col['x0'] - prev_col['x1']) * 0.2
                    
                    if ci == len(merged_cols) - 1:
                        x_end = 9999
                    else:
                        # End of this column is influenced by the next column's start
                        next_col = merged_cols[ci + 1]
                        x_end = col['x1'] + (next_col['x0'] - col['x1']) * 0.2
                    
                    col_ranges.append((x_start, x_end))
                
                self.log(f"  -> Detected columns: {col_names}")
                
                # Step 3: Extract all data words and assign to columns by x-position
                all_word_rows = []
                
                for page_idx, page in enumerate(pdf.pages):
                    words = page.extract_words()
                    
                    # Skip words above header on the header page
                    if page_idx == header_page_idx and header_top is not None:
                        words = [w for w in words if w['top'] > header_top + 10]
                    elif page_idx < header_page_idx:
                        continue
                    
                    # Group words by y position (same visual line)
                    y_groups = {}
                    for w in words:
                        y_key = round(w['top'] / 4) * 4
                        if y_key not in y_groups:
                            y_groups[y_key] = []
                        y_groups[y_key].append(w)
                    
                    for y_key in sorted(y_groups.keys()):
                        all_word_rows.append(y_groups[y_key])
                
                # Step 4: Build structured rows from word positions
                raw_rows = []
                for words_in_line in all_word_rows:
                    row = [''] * len(col_names)
                    for w in sorted(words_in_line, key=lambda x: x['x0']):
                        word_center = (w['x0'] + w['x1']) / 2
                        best_col = 0
                        for ci, (x_start, x_end) in enumerate(col_ranges):
                            if x_start <= word_center <= x_end:
                                best_col = ci
                                break
                        if row[best_col]:
                            row[best_col] += ' ' + w['text']
                        else:
                            row[best_col] = w['text']
                    
                    # Skip boilerplate rows
                    row_text = ' '.join(row)
                    if is_skip_line(row_text):
                        continue
                    if not any(cell.strip() for cell in row):
                        continue
                    
                    # Cleanup: If numeric columns (In/Out/Balance/Debit/Credit) contain text with letters, 
                    # move it to the description/transaction column
                    num_col_indices = []
                    desc_col_idx = -1
                    for ci, name in enumerate(col_names):
                        lname = name.lower()
                        if any(kw in lname for kw in ['in', 'out', 'balance', 'debit', 'credit', 'amount']):
                            num_col_indices.append(ci)
                        if any(kw in lname for kw in ['transaction', 'description', 'details']):
                            desc_col_idx = ci
                    
                    if desc_col_idx != -1:
                        for nci in num_col_indices:
                            val = row[nci].strip()
                            # If it contains letters (not just currency symbols or Cr/Dr), it's probably description spillover
                            clean_val = val.lower().replace('cr', '').replace('dr', '').replace('aed', '').replace('usd', '').strip()
                            if any('a' <= c <= 'z' for c in clean_val):
                                if row[desc_col_idx]:
                                    row[desc_col_idx] += ' ' + val
                                else:
                                    row[desc_col_idx] = val
                                row[nci] = ''
                    
                    raw_rows.append(row)
            
            # Step 5: Merge multi-line transactions
            # A new transaction starts when the Date column has a date
            self.log(f"  -> Total raw rows after filtering: {len(raw_rows)}")
            if raw_rows:
                self.log(f"  -> First raw row: {raw_rows[0]}")
                
            transactions = []
            current_txn = None
            
            for row in raw_rows:
                # Fix for split dates (e.g., "30 Apr," in col 0 and "2025 ..." in col 1)
                if len(row) > 1:
                    combined = (row[0] + ' ' + row[1]).strip()
                    match = date_pattern.search(combined)
                    # If the combined text starts with a date that wasn't in row[0] alone
                    if match and not date_pattern.search(row[0]):
                        full_date = match.group(0)
                        # Move the date part to row[0] and keep the rest in row[1]
                        remainder = combined[match.end():].strip()
                        row[0] = full_date
                        row[1] = remainder
                
                has_date = date_pattern.search(row[0]) if row[0] else False
                
                if has_date:
                    if current_txn:
                        transactions.append(current_txn)
                    current_txn = [list(row)]
                else:
                    if current_txn:
                        current_txn.append(list(row))
            
            if current_txn:
                transactions.append(current_txn)
                
            self.log(f"  -> Transactions grouped: {len(transactions)}")
            
            if not transactions:
                self.log(f"  -> No transactions found in {pdf_file}.")
            else:
                # Merge multi-row transactions into single rows
                output_data = [col_names]
                
                for txn_rows in transactions:
                    merged = [''] * len(col_names)
                    for row in txn_rows:
                        for ci, cell in enumerate(row):
                            if cell.strip():
                                if merged[ci]:
                                    merged[ci] += '\n' + cell.strip()
                                else:
                                    merged[ci] = cell.strip()
                    output_data.append(merged)
                
                df = pd.DataFrame(output_data)
                base_name = os.path.splitext(pdf_file)[0]
                excel_path = os.path.join(output_folder, f"{base_name}.xlsx")
                
                save_styled_excel(df, excel_path)
                self.log(f"  -> Extracted {len(transactions)} transactions. Saved to {excel_path}")
                    
        except Exception as e:
            import traceback
            self.log(f"  -> Error processing {pdf_file}: {str(e)}")
            self.log(traceback.format_exc())
            
        self.log("\nExtraction complete! File saved in:")
        self.log(output_folder)
        
        QMessageBox.information(self, "Success", "Extraction complete!\nFile saved in the extracted_output folder.")
        
        self.extract_btn.setEnabled(True)
        self.select_file_btn.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    
    window = App()
    window.show()
    sys.exit(app.exec())

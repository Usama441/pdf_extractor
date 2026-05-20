import os
import re
from dataclasses import dataclass

import pandas as pd
import pdfplumber


class ExtractionError(Exception):
    def __init__(self, message, error_code):
        super().__init__(message)
        self.error_code = error_code


class PasswordRequiredError(ExtractionError):
    pass


class UnsupportedStatementError(ExtractionError):
    pass


DATE_PATTERN = re.compile(
    r"\b(?:\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|"
    r"\d{1,2}[A-Z]{3}\d{2}|\d{1,2}\s+[A-Za-z]{3},?\s+\d{4})\b"
)

REFERENCE_PATTERN = re.compile(
    r"^(?:"
    r"[A-Z0-9]{6,}"
    r"|[A-Z0-9]{4,}\s+[A-Z0-9]{3,}"
    r"|[A-Z]{2,}[-/]?[A-Z0-9]{4,}"
    r")$"
)

AMOUNT_PATTERN = re.compile(r"^(?:[A-Z]{3}\s*)?[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?(?:\s*(?:CR|DR))?$", re.I)

HEADER_KEYWORDS = [
    "date",
    "description",
    "balance",
    "amount",
    "withdrawal",
    "deposit",
    "credit",
    "debit",
    "transaction",
    "type",
    "ref.",
    "reference",
    "debits",
    "credits",
    "value date",
    "chq/ref",
    "currency",
    "money out",
    "money in",
    "status",
]

KNOWN_MULTI_WORD_HEADERS = {
    "reference number": "Reference Number",
    "account balance": "Account Balance",
    "ref. number": "Ref. Number",
    "payments in": "Payments In",
    "payments out": "Payments Out",
    "value date": "Value Date",
    "chq/ref no.": "Chq/Ref No.",
    "chq/ref": "Chq/Ref No.",
    "running balance": "Running Balance",
    "money out": "Money Out",
    "money in": "Money In",
    "date (utc)": "Date",
}

SKIP_KEYWORDS = [
    "your bank statement",
    "account summary",
    "opening balance",
    "closing balance",
    "total deposits",
    "total withdrawals",
    "interest rate",
    "overdue charges",
    "interest type",
    "please see",
    "important information",
    "the national bank",
    '"rakbank"',
    "central bank of",
    "page [",
    "page ",
    "disclaimer",
    "end of day",
    "summary ",
    "payments in",
    "payments out",
    "get in touch",
    "for queries",
    "for complaints",
    "call 04",
    "we aim",
    "this is a digitally",
    "accrued interest",
    "account number:",
    "branch:",
    "currency:",
    "iban:",
    "account type:",
    "statement period:",
    "date issued:",
    "24hr customer service",
    "www.starling",
    "sort code:",
    "account name:",
    "financial services",
    "your deposit",
    "© 2025",
    "wio bank",
    "this bank is regulated",
    "please review",
    "within 30 days",
    "correct (subject",
    "to raise a complaint",
    "summary of accounts",
    "account statement",
    "account holder",
    "the items and balance",
    "verified. report any",
    "of the statement date",
    "accurate.",
    "all charges, terms",
    "please note that for foreign",
    "indicative only",
    "mashreqbank psc",
    "it is the policy",
    "full compliance",
    "with iran, syria",
    "dear customer",
    "statement for period",
    "customer number",
    "account currency",
    "account number ",
    "يكنبلا",
    "يراجلا",
    "خيراتلا",
    "فصولا",
    "كيشلا",
    "بحسلا",
    "ةعيدولا",
    "ديصرلا",
    "باسحلا",
    "ةدئافلا",
    "ةرخأتملا",
    "ةفاضملا",
    ".ةمهم",
    "ينطولا",
    "يزكرملا",
    "تامولعم",
    "فشكلا",
    "رادصإلا",
    "account balance",
    "carried forward",
    "is registered in england",
    "prudential regulation authority",
    "financial conduct authority",
    "london fruit and wool exchange",
    "fscs website",
    "information sheet and exclusions",
    "authorised by the",
    "regulated by the",
    "starling bank limited",
    "duval square",
    "e1 6pw",
    "number 730166",
    "list which are",
    "available in the app and",
    "on our website.",
    "further details can also be found",
    "refer to the fscs website",
    "we charge",
    "interest each day you",
    "date range applicable",
    "unlimited 0.00%",
    "less than",
    "interest rates",
    "overdraft",
    "united arab",
    "emirates p.o box",
    "digital stamp",
    "does not require signature",
    "description (incl. vat)",
    "amount balance",
    "ref. number",
    "tax registration number",
    "brought forward",
    "need help?",
    "following channels",
    "personal banking",
    "business banking",
    "private banking",
    "dedicated relationship",
    "online banking",
    "future reference",
    "dispute resolution",
    "sanadak.ae",
    "centralbank.ae",
    "complaint with the bank",
    "contact us",
    "customer service",
    "registered details:",
    "paid up capital",
    "commercial registration",
    "head office:",
    "p.j.s.c",
    "licensed by the central bank",
    "emirates nbd bank",
    "phone banking:",
    "email at",
    "complaint-process",
    "customersupport@",
    "www.emiratesnbd.com",
    "visiting any",
    "nbd branch",
    "emirates nbd branch",
    "800 54",
    "800456",
    "p.o. box",
    "uae",
    "date description",
    "debits credits",
    "credits balance",
    "description debits",
    "end of statement",
    "total debits",
    "total credits",
    "total amount",
    "total",
    "generated on",
    "filters and search applied",
    "transaction statement",
    "report lost or",
    "scan the qr code",
    "revolut ltd",
    "transfer of funds",
    "regulations 2017",
    "information on the payer",
    "accounts:",
]


@dataclass
class ExtractionResult:
    columns: list[str]
    rows: list[list[str]]
    transaction_count: int
    source_filename: str

    def to_dict(self):
        return {
            "columns": self.columns,
            "rows": self.rows,
            "transaction_count": self.transaction_count,
            "source_filename": self.source_filename,
        }


def extract_pdf_to_excel(pdf_path, excel_path, password=None):
    validate_pdf_access(pdf_path, password=password)

    with pdfplumber.open(pdf_path, password=password) as pdf:
        all_lines = extract_lines(pdf)
        if not all_lines:
            raise UnsupportedStatementError("No text could be extracted from this PDF.", "no_text_found")

        header_line = detect_header_line(all_lines)
        if not header_line:
            raise UnsupportedStatementError("Could not detect a transaction header in this PDF.", "header_not_found")

        columns, column_ranges, header_page_idx, header_top = detect_column_layout(pdf)
        raw_rows = extract_rows(pdf, columns, column_ranges, header_page_idx, header_top)
        transactions = merge_transactions(raw_rows)

    if not transactions:
        raise UnsupportedStatementError("No transactions were detected in this PDF.", "no_transactions_found")

    save_styled_excel(columns, transactions, excel_path)

    return ExtractionResult(
      columns=columns,
      rows=transactions,
      transaction_count=len(transactions),
      source_filename=os.path.basename(pdf_path),
    )


def validate_pdf_access(pdf_path, password=None):
    try:
        with pdfplumber.open(pdf_path, password=password):
            return
    except Exception as exc:  # noqa: BLE001
        error_msg = str(exc).lower()
        error_type = type(exc).__name__.lower()

        if any(keyword in error_msg for keyword in ["password", "authenticate", "pdfminer", "encrypted"]) or any(
            keyword in error_type for keyword in ["password", "pdfminer", "crypt"]
        ):
            if password:
                raise PasswordRequiredError(
                    "The provided PDF password was invalid. Retry with the correct password.",
                    "password_required",
                ) from exc

            raise PasswordRequiredError(
                "The uploaded PDF is password protected. Provide a valid password and retry.",
                "password_required",
            ) from exc


def extract_lines(pdf):
    lines = []
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue
        lines.extend(line.strip() for line in text.split("\n"))
    return lines


def detect_header_line(lines):
    for line in lines:
        lower = line.lower().strip()
        if "date" not in lower:
            continue

        matches = sum(1 for keyword in HEADER_KEYWORDS if keyword in lower)
        if matches >= 2:
            return line

    return None


def detect_column_layout(pdf):
    for page_idx, page in enumerate(pdf.pages):
        words = page.extract_words()
        groups = group_words_by_line(words)
        sorted_y_keys = sorted(groups.keys())

        for y_key in sorted_y_keys:
            line_words = groups[y_key]
            line_text = " ".join(word["text"].lower() for word in line_words)
            matches = sum(1 for keyword in HEADER_KEYWORDS if keyword in line_text)

            if matches >= 3 and "date" in line_text:
                header_top = line_words[0]["top"]
                header_words = []

                for adjacent_y in sorted_y_keys:
                    if abs(adjacent_y - y_key) <= 12:
                        for word in groups[adjacent_y]:
                            word_text = word["text"].lower()
                            if word_text in HEADER_KEYWORDS or word_text in (
                                "in",
                                "out",
                                "number",
                                "cheque",
                                "ref.",
                                "reference",
                                "account",
                                "no.",
                                "value",
                                "running",
                            ):
                                header_words.append(word)

                        if adjacent_y > y_key:
                            header_top = max(header_top, max(word["top"] for word in groups[adjacent_y]))

                merged_columns = merge_header_words(header_words)
                columns = [column["text"] for column in merged_columns]
                return columns, build_column_ranges(merged_columns), page_idx, header_top

    raise UnsupportedStatementError(
        "Could not locate transaction column positions in this PDF.",
        "column_positions_not_found",
    )


def merge_header_words(header_words):
    header_words = sorted(header_words, key=lambda word: word["x0"])
    columns = []
    index = 0

    while index < len(header_words):
        word = header_words[index]
        if any(ord(char) > 127 for char in word["text"]):
            index += 1
            continue

        if index + 2 < len(header_words):
            second = header_words[index + 1]
            third = header_words[index + 2]
            three_word = f"{word['text']} {second['text']} {third['text']}".lower()
            if three_word in KNOWN_MULTI_WORD_HEADERS:
                columns.append({"text": KNOWN_MULTI_WORD_HEADERS[three_word], "x0": word["x0"], "x1": third["x1"]})
                index += 3
                continue

        if index + 1 < len(header_words):
            second = header_words[index + 1]
            two_word = f"{word['text']} {second['text']}".lower()
            if two_word in KNOWN_MULTI_WORD_HEADERS:
                columns.append({"text": KNOWN_MULTI_WORD_HEADERS[two_word], "x0": word["x0"], "x1": second["x1"]})
                index += 2
                continue

        columns.append({"text": word["text"].capitalize(), "x0": word["x0"], "x1": word["x1"]})
        index += 1

    return columns


def build_column_ranges(columns):
    ranges = []
    for index, column in enumerate(columns):
        if index == 0:
            x_start = 0
        else:
            previous = columns[index - 1]
            x_start = previous["x1"] + (column["x0"] - previous["x1"]) * 0.2

        if index == len(columns) - 1:
            x_end = 9999
        else:
            following = columns[index + 1]
            bias = 0.8 if column["text"].lower() in ("date", "description") else 0.5
            if column["text"].lower() in ("status", "account"):
                bias = 0.2
            x_end = column["x1"] + (following["x0"] - column["x1"]) * bias

        ranges.append((x_start, x_end))

    return ranges


def extract_rows(pdf, columns, column_ranges, header_page_idx, header_top):
    raw_rows = []

    for page_idx, page in enumerate(pdf.pages):
        words = page.extract_words()
        if page_idx == header_page_idx:
            words = [word for word in words if word["top"] > header_top + 10]
        elif page_idx < header_page_idx:
            continue

        grouped_lines = group_words_by_line(words)
        page_date_seen = page_idx == header_page_idx
        for y_key in sorted(grouped_lines.keys()):
            row = assign_words_to_columns(columns, column_ranges, grouped_lines[y_key])
            row_text = " ".join(row)

            if is_skip_line(row_text) or not any(cell.strip() for cell in row):
                continue

            if DATE_PATTERN.search(f"{row[0]} {row[1]}".strip()):
                page_date_seen = True
            elif page_idx > header_page_idx and not page_date_seen:
                continue

            raw_rows.append(clean_numeric_spillover(columns, row))

    return raw_rows


def assign_words_to_columns(columns, column_ranges, words):
    row = [""] * len(columns)

    for word in sorted(words, key=lambda item: item["x0"]):
      center = (word["x0"] + word["x1"]) / 2
      target_index = 0
      for index, (x_start, x_end) in enumerate(column_ranges):
          if x_start <= center <= x_end:
              target_index = index
              break

      row[target_index] = f"{row[target_index]} {word['text']}".strip()

    return row


def merge_transactions(raw_rows):
    transactions = []
    current_transaction = None

    for row in raw_rows:
        row = clean_reference_spillover(row)
        if len(row) > 1:
            combined = f"{row[0]} {row[1]}".strip()
            match = DATE_PATTERN.search(combined)
            if match:
                row[0] = match.group(0)
                row[1] = f"{combined[:match.start()]} {combined[match.end():]}".strip()

        has_date = DATE_PATTERN.search(row[0]) if row[0] else False

        if has_date:
            if current_transaction:
                transactions.append(current_transaction)
            current_transaction = [list(row)]
        elif current_transaction:
            current_transaction.append(list(row))

    if current_transaction:
        transactions.append(current_transaction)

    merged = []
    for rows in transactions:
        combined_row = [""] * len(rows[0])
        for row in rows:
            for index, cell in enumerate(row):
                if cell.strip():
                    combined_row[index] = f"{combined_row[index]}\n{cell.strip()}".strip()
        merged.append(combined_row)

    return deduplicate_transactions(merged)


def clean_reference_spillover(row):
    if len(row) < 3:
        return row

    has_date = bool(DATE_PATTERN.search(row[0])) if row[0] else False
    if has_date:
        return row

    reference = row[2].strip()
    if not reference or is_reference_like(reference) or is_amount_like(reference):
        return row

    row = list(row)
    row[1] = f"{row[1]} {reference}".strip()
    row[2] = ""
    return row


def deduplicate_transactions(transactions):
    deduplicated = []
    seen = set()

    for transaction in transactions:
        signature = transaction_signature(transaction)
        if signature in seen:
            continue

        seen.add(signature)
        deduplicated.append(transaction)

    return deduplicated


def transaction_signature(transaction):
    return tuple(normalize_signature_cell(cell) for cell in transaction)


def normalize_signature_cell(value):
    return re.sub(r"\s+", " ", str(value).strip().lower())


def is_reference_like(value):
    normalized = re.sub(r"\s+", " ", value.strip().upper())
    if not normalized:
        return False

    return bool(REFERENCE_PATTERN.match(normalized))


def is_amount_like(value):
    normalized = value.strip().replace(",", "")
    return bool(AMOUNT_PATTERN.match(normalized))


def save_styled_excel(columns, transactions, excel_path):
    dataframe = pd.DataFrame([columns] + transactions)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, header=False, sheet_name="Transactions")
        worksheet = writer.sheets["Transactions"]

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


def clean_numeric_spillover(columns, row):
    numeric_column_indexes = []
    description_index = -1

    for index, name in enumerate(columns):
        lower_name = name.lower()
        if any(keyword in lower_name for keyword in ["in", "out", "balance", "debit", "credit", "amount"]):
            numeric_column_indexes.append(index)
        if any(keyword in lower_name for keyword in ["transaction", "description", "details"]):
            description_index = index

    if description_index == -1:
        return row

    for index in numeric_column_indexes:
        value = row[index].strip()
        cleaned = value.lower().replace("cr", "").replace("dr", "").replace("aed", "").replace("usd", "").strip()
        if any("a" <= char <= "z" for char in cleaned):
            row[description_index] = f"{row[description_index]} {value}".strip()
            row[index] = ""

    return row


def group_words_by_line(words):
    groups = {}
    for word in words:
        key = round(word["top"] / 4) * 4
        groups.setdefault(key, []).append(word)
    return groups


def is_skip_line(line):
    lower = line.lower().strip()
    if not lower:
        return True

    for keyword in SKIP_KEYWORDS:
        if keyword in lower:
            return True

    if any("\u0600" <= char <= "\u06FF" for char in line):
        return True

    if re.search(r"page\s*\d+", lower):
        return True

    letters = sum(1 for char in lower if "a" <= char <= "z")
    digits = sum(1 for char in lower if "0" <= char <= "9")
    if letters == 0 and digits == 0 and len(lower) > 2:
        return True

    if re.search(r"\+\d+|\d{3}-\d{4}|\d{4}\s\d{4}|www\.|http", lower):
        return True

    return lower in ["date", "description", "debits", "credits", "balance", "total"]
